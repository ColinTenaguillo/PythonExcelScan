#!/usr/bin/env python
# -*-coding:Utf-8 -*

""".py: Scan et analyse les tickets du fichier excel."""

import openpyxl
import mef  # where i create my class service and his functions

# settings for the xlsx that we are going to read
nomxlsx_filepath = "Glpi_tracking/glpi_tracking2.xlsx"  # filepath of the xlsx
wb = openpyxl.load_workbook(nomxlsx_filepath)
sheet = wb.sheetnames
ws = wb[sheet[0]]
row_count = ws.max_row
column_count = ws.max_column
rowrange = "A1:M" + str(row_count)
# ---------------------------------------------------
# settings for the xlsx we are going to write in
xlsx_results_filepath = "Resultats/Results2.xlsx"
wb_results = openpyxl.load_workbook(xlsx_results_filepath)
sheet_results = wb_results.active
sheet_results.cell(row=1, column=1).value = "Nom ticket"
sheet_results.cell(row=1, column=2).value = "Nombres de tickets"
sheet_results.cell(row=1, column=3).value = "Temps de traitement"
# ----------------------------------------------------
services = [
    mef.Service("check_dp"),
    mef.Service("LOAD"),
    mef.Service("Load"),
    mef.Service("vtom-jobs"),
    mef.Service("SERVICES_AUTO"),
    mef.Service("NTP"),
    mef.Service("EVENTLOG-SYSTEM"),
    mef.Service("REBOOT"),
    mef.Service("DISK"),
    mef.Service("SM37-AbortedJobs"),
    mef.Service("PAGEFILE"),
    mef.Service("LOCKLONG"),
    mef.Service("BATCH_ECHEC"),
    mef.Service("CPU"),
    mef.Service("FS_"),
    mef.Service("TALEND"),
    mef.Service("JVM_HEALTH"),
    mef.Service("HARDWARE"),
    mef.Service("DEADLOCK"),
    mef.Service("ST22 - Dumps ABAP"),
]

i = 1
for _row in ws[rowrange]:
    # Column C is where the ticket name is stored
    C = str(ws.cell(row=i, column=3).value)
    # If a cell is empty it returns "NULL" that occurs a ValueError
    try:
        # Column M is the data i want to analyse(ticket time)
        M = int(ws.cell(row=i, column=13).value)
    except ValueError:
        M = 0
    for Service in services:
        Service.ajoute(C, M)
    i += 1


i = 2  # we put i at 2 cause the first row is used for column name
for Service in services:
    Service.print_excel(i, sheet_results)
    i += 1

wb_results.save(xlsx_results_filepath)
